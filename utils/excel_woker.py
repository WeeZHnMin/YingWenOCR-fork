from PyQt5.QtCore import QThread, pyqtSignal, QBuffer, QIODevice
from PyQt5.QtGui import QPixmap
from PIL import Image
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import os

def save_to_excel(data_list, base_path, file_name="ocr_results"):
    """
    将OCR结果保存到Excel文件
    
    Args:
        data_list: 包含 (pil_image, character, confidence) 元组的列表
        base_path: 基础路径
        file_name: 文件名（不含扩展名）
    """
    # 创建输出文件夹
    output_folder = os.path.join(base_path, "output")
    os.makedirs(output_folder, exist_ok=True)
    
    file_path = os.path.join(output_folder, f"{file_name}.xlsx")
    
    # 检查文件是否存在，决定是加载还是创建
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        # 获取下一行的行号
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "OCR Results"
        # 写入表头
        ws.append(["Row", "Image", "Character", "Confidence"])
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        start_row = 2  # 从第2行开始（第1行是表头）
    
    # 处理每个数据项
    for index, (pil_image, character, confidence) in enumerate(data_list):
        current_row = start_row + index
        
        # 添加行号和文本数据
        ws.cell(row=current_row, column=1, value=current_row - 1)  # 行号
        ws.cell(row=current_row, column=3, value=character)        # 字符
        ws.cell(row=current_row, column=4, value=confidence)       # 置信度
        
        # 处理图片
        try:
            # 将PIL Image转换为字节流
            img_buffer = BytesIO()
            
            # 确保图片格式兼容
            if pil_image.mode in ('RGBA', 'LA', 'P'):
                pil_image = pil_image.convert('RGB')
            
            # 调整图片尺寸（可选）
            max_size = (120, 80)  # 适合单个字符的尺寸
            if pil_image.size[0] > max_size[0] or pil_image.size[1] > max_size[1]:
                pil_image.thumbnail(max_size, Image.Resampling.LANCZOS)
            
            # 保存到字节流
            pil_image.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # 创建Excel图片对象
            excel_img = ExcelImage(img_buffer)
            excel_img.anchor = f'B{current_row}'  # 锚定到B列对应行
            
            # 添加图片到工作表
            ws.add_image(excel_img)
            
            # 调整行高以适应图片
            ws.row_dimensions[current_row].height = 60
            
        except Exception as e:
            print(f"添加第 {index + 1} 个图片失败: {e}")
            ws.cell(row=current_row, column=2, value="[图片添加失败]")
    
    # 保存文件
    try:
        wb.save(file_path)
        print(f"成功保存 {len(data_list)} 条数据到: {file_path}")
    except Exception as e:
        print(f"保存文件失败: {e}")


class SaveExcelWorker(QThread):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    
    def __init__(self, table_data):
        super().__init__()
        self.table_data = table_data
    
    def qpixmap_to_pil(self, qpixmap):
        if qpixmap is None or qpixmap.isNull():
            return None
            
        buffer = QBuffer()
        buffer.open(QIODevice.WriteOnly)
        qpixmap.save(buffer, "PNG")
        buffer.close()
        
        buffer.open(QIODevice.ReadOnly)
        data = buffer.data()
        bytes_io = BytesIO(data.data())
        pil_image = Image.open(bytes_io)
        buffer.close()
        return pil_image.copy()
    
    def run(self):
        try:
            data_list = []
            for thumbnail_viewer, character, confidence in self.table_data:
                if hasattr(thumbnail_viewer, 'original_pixmap'):
                    pixmap = thumbnail_viewer.original_pixmap
                    pil_image = self.qpixmap_to_pil(pixmap)
                    if pil_image:
                        data_list.append((pil_image, character, confidence))
            
            if data_list:
                save_to_excel(data_list, ".")
            
            self.finished.emit()
        except Exception as e:
            self.error.emit(str(e))