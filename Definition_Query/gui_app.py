import tkinter as tk
from tkinter import filedialog, messagebox
from threading import Thread
from PIL import Image
import requests
from urllib.parse import unquote
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook, load_workbook
from logo import get_icon_photo_from_base64

# =================== 原有函数保留 ===================

def make_simple_request(word):
    url = "https://ytenx.org/zim"
    params = {
        'dzih': word,
        'dzyen': '1',
        'jtkb': '1',
        'jtkd': '1',
        'jtdt': '1',
        'jtgt': '1'
    }
    headers = {
        'user-agent': 'Mozilla/5.0',
        'referer': 'https://ytenx.org/'
    }
    response = requests.get(url, params=params, headers=headers)
    return response

def parse_yonh_content(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    try:
        yonh_tags = soup.find_all('p', class_='yonh')
        element = yonh_tags[0]
        res = element.get_text(strip=True).split('(')[0]
        return res
    except:
        return None

def read_excel(input_file, column_letters=["C", "D"]):
    if not os.path.exists(input_file):
        return None
    wb = load_workbook(input_file)
    ws = wb.active
    data = []
    col_indices = [ord(letter.upper()) - ord('A') + 1 for letter in column_letters]
    for row in ws.iter_rows(min_row=2):
        row_data = [row[i - 1].value for i in col_indices]
        data.append(row_data)
    return data

def save_excel(output_file, datas: list):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = '置信度'
    ws['B1'] = '字'
    for i in range(6):
        col_letter = chr(ord('C') + i)
        ws[f'{col_letter}1'] = chr(ord('A') + i)
    for data in datas:
        ws.append(list(data))
    wb.save(output_file)

# =================== 主程序函数 ===================

def run_task(input_file, output_file, log_func):
    data = read_excel(input_file)
    if not data:
        log_func("❌ 输入文件为空或不存在！")
        return

    all_data = []
    for line in data:
        char = line[0]
        response = make_simple_request(char)
        cur_lst = [line[1], char]
        if response.status_code == 200:
            content = unquote(response.text)
            res = parse_yonh_content(content)
            if res:
                log_func(f"✅ 字:{char}, 释义:{res}")
                cur_lst += list(res)
            else:
                log_func(f"❌ 字:{char}, 释义:未找到")
            all_data.append(cur_lst)

    save_excel(output_file, all_data)
    log_func("✅ 已保存结果到：" + output_file)

# =================== UI 界面部分 ===================

def start_gui():
    def browse_input_file():
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        input_entry.delete(0, tk.END)
        input_entry.insert(0, path)

    def run():
        input_file = input_entry.get()
        output_file = output_entry.get()
        if not input_file or not output_file:
            messagebox.showwarning("警告", "请输入输入和输出路径")
            return
        log_box.delete(1.0, tk.END)
        log("🚀 开始处理，请等待...\n")

        # 使用线程防止界面卡死
        thread = Thread(target=run_task, args=(input_file, output_file, log))
        thread.start()

    def log(text):
        log_box.insert(tk.END, text + '\n')
        log_box.see(tk.END)  # 滚动到底部

    # GUI 布局
    window = tk.Tk()
    window.title("古文字释义处理器")
    icon_photo = get_icon_photo_from_base64()
    window.iconphoto(False, icon_photo)  # 设置窗口图标
    window.geometry("700x500")

    # 输入文件选择
    tk.Label(window, text="输入Excel文件：").pack(anchor="w", padx=10, pady=5)
    frame_input = tk.Frame(window)
    frame_input.pack(fill="x", padx=10)
    input_entry = tk.Entry(frame_input, width=60)
    input_entry.pack(side="left", fill="x", expand=True)
    tk.Button(frame_input, text="浏览", command=browse_input_file).pack(side="right")

    # 输出文件选择
    tk.Label(window, text="输出Excel文件：").pack(anchor="w", padx=10, pady=5)
    frame_output = tk.Frame(window)
    frame_output.pack(fill="x", padx=10)
    output_entry = tk.Entry(frame_output, width=60)
    output_entry.pack(side="left", fill="x", expand=True)

    def browse_output_file():
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="选择保存文件路径"
        )
        if path:
            output_entry.delete(0, tk.END)
            output_entry.insert(0, path)

    tk.Button(frame_output, text="浏览", command=browse_output_file).pack(side="right")

    # 执行按钮
    tk.Button(window, text="执行处理", command=run, bg="green", fg="white").pack(pady=10)

    # 日志输出框
    log_box = tk.Text(window, height=20)
    log_box.pack(fill="both", expand=True, padx=10, pady=5)

    window.mainloop()

# 启动程序
if __name__ == "__main__":
    start_gui()
